#!/usr/bin/env python
# coding: utf-8

# In[ ]:


# Importem les llibreries necessàries
from openpyxl import load_workbook
from datetime import datetime
from bs4 import BeautifulSoup
import pandas as pd
import requests
import time

# Definim el número màxim de pàgines a consultar
max_pages = 5

# Definim el temps a esperar entre cada connexió a les diferents pàgines
sleeping_time = 15

# Fem una llista on emmagatzemar les URLs dels pisos ofertats
urls_pisos = []

# Accedim a les N pàgines que haguem definit
for page in range(1, max_pages + 1):
    
    # Definim la URL de les finques d'on volem obtenir la informació dels pisos
    url = f"https://www.finquessip.com/index.php?limtipos=21299%2C399%2C499%2C199%2C2999%2C3399&areas=11925_idealista%2C14636_idealista%2C14427_idealista&buscador=1&idio=1&pag={page}#modulo-paginacion"

    try:
        # Fem la petició a la pàgina
        response = requests.get(url)
        
        # Verifiquem si ha sigut exitosa
        if response.status_code == 200:
            print(f"[SUCCESS] Accedint correctament a la pàgina {page} / {max_pages} ...")
            
            # Capturem el contignut HTML amb BeautifulSoup
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Busquem totes les etiquetes 'a' de la classe que conté les URLs de cada pis
            links = soup.find_all('a', class_='irAfichaPropiedad enviarPostHog')
            
            # Obtenim les URLs de cada pis i les guardem a la variable 'urls_pisos'
            for link in links:
                href = link.get('href')
                if href:
                    urls_pisos.append('https://www.finquessip.com/' + href)
                    
        else:
            # Si no es pot accedir a la pàgina mostrem l'error per pantalla
            print(f"[ERROR] No es pot accedir a la pàgina {page} / {max_pages}, estat: {response.status_code}")
            break
        
        # Esperem el temps d'espera entre connexions per no saturar la web
        time.sleep(sleeping_time)
    
    except Exception as e:
        
        # Si salta alguna excepció la mostrem per pantalla
        print(f"[ERROR] No es pot accedir a la pàgina {page} / {max_pages}, excepció: {str(e)}")
        
        break

# Eliminem la meitat dels registres ja que la informació està duplicada       
urls_pisos = urls_pisos[1::2]        

# Guardem en una variable el total de URLs de pisos obtingudes
total_urls = len(urls_pisos)

# Mostrem per pantalla la quantitat total de URLs de pisos obtingudes
print(f"\n S'han pogut capturar {total_urls} URLs de pisos diferents")


# Generem un dataframe buit amb les columnes dels diferents atributs que volem capturar de cada URL
columnes = ['Estat', 'Data_captura', 'Plataforma', 'Referencia', 'Data_actualització', 'Diferencia_temps', 'Gap_1', 'Preu', 'Comunitat', 'Habitatge', 'Conservacio', 'Any', 'Ciutat', 'Barri', 'm2_constr', 'm2_utils', 'Planta', 'Ascensor', 'Habitacions', 'Banys', 'Orientacio', 'Jardi', 'Terrassa', 'Piscina', 'Garatge', 'Traster', 'URL']
df_pisos = pd.DataFrame(columns=columnes)

# Definim una funció per extreure el text de dins una etiqueta HTML
def extraer_valor(soup, clase_caracteristica):
    elemento = soup.find('span', string=clase_caracteristica)
    if elemento:
        valor = elemento.find_next('span', class_='valor')
        return valor.get_text(strip=True) if valor else None
    return None

# Accedim a cada una de les URLs individuals de pisos capturades anteriorment
for i, url in enumerate(urls_pisos, start=1):
    
    try:
        # Fem la petició a la pàgina
        response = requests.get(url)
        
        # Verifiquem si ha sigut exitosa
        if response.status_code == 200:
            print(f"[SUCCESS] Accedint correctament a la URL {i} / {total_urls} : {url}")
            
            # Capturem el contignut HTML amb BeautifulSoup
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Extraiem el preu
            preu = soup.find('div', class_='fichapropiedad-precio').get_text(strip=True)
            
            # Extraiem el llistat de característiques que no estan indexades com la resta
            etiquetas = soup.find_all('b', class_='etiqueta')
            textos_etiqueta = [etiqueta.get_text().strip() for etiqueta in etiquetas]
            
            # Fem una llista dels atributs d'interès a buscar dins de la llista anterior
            ascensor = 'No'
            jardi = 'No'
            terrassa = 'No'
            piscina = 'No'
            garatge = 'No'
            traster = 'No'

            # Accedim a la llista en busca dels atributs
            for texto in textos_etiqueta:
                    
                if 'ascensor' in texto.lower(): 
                    ascensor = 'Si'
                    
                if 'jardín' in texto.lower(): 
                    jardi = 'Si'
                    
                if 'terraza' in texto.lower(): 
                    terrassa = 'Si'
                                
                if 'piscina' in texto.lower(): 
                    piscina = 'Si'
                
                if 'parking' in texto.lower(): 
                    garatge = 'Si'
                    
                if 'trastero' in texto.lower(): 
                    traster = 'Si'
        
            # Extraiem la resta de característiques que estan indexades correctament
            referencia = extraer_valor(soup, "Referencia ")
            comunitat = extraer_valor(soup, "Gastos Comunidad")
            habitatge = extraer_valor(soup, "Tipo de Propiedad")
            conservacio = extraer_valor(soup, "Conservación")
            anys = extraer_valor(soup, "Antigüedad")
            zona = extraer_valor(soup, "Zona / Ciudad ")
            barri, ciutat = zona.split(' / ')
            m2_constr = extraer_valor(soup, "Superficie Construida")
            m2_utils = extraer_valor(soup, "Superficie Útil")
            planta = extraer_valor(soup, "Planta ")
            habitacions = extraer_valor(soup, "Habitaciones")
            banys = extraer_valor(soup, "Baños")
            orientacio = extraer_valor(soup, "Orientación")
           
            # Afegim tota la informació capturada al dataframe creat anteriorment
            df_pisos = df_pisos.append({'Estat': 'Actiu',
                                        'Data_captura':datetime.today().strftime('%Y-%m-%d'),
                                        'Plataforma': 'Finques SIP',
                                        'Referencia': referencia,
                                        'Data_actualització':datetime.today().strftime('%Y-%m-%d'),
                                        'Diferencia_temps': 0,
                                        'Gap_1': '',
                                        'Preu': preu,
                                        'Comunitat': comunitat,
                                        'Habitatge': habitatge,
                                        'Conservacio': conservacio, 
                                        'Any': anys,
                                        'Ciutat': ciutat,
                                        'Barri': barri,
                                        'm2_constr': m2_constr,
                                        'm2_utils': m2_utils,
                                        'Planta': planta,
                                        'Ascensor': ascensor,
                                        'Habitacions': habitacions,
                                        'Banys': banys,
                                        'Orientacio': orientacio,
                                        'Jardi': jardi,
                                        'Terrassa': terrassa,
                                        'Piscina': piscina,
                                        'Garatge': garatge,
                                        'Traster': traster,
                                        'URL': url
                                       }, ignore_index=True)
        
        else:
            print(f"[ERROR] No es pot accedir a la URL {i} / {total_urls} : {url}, estat: {response.status_code}")

        # Esperem el temps d'espera entre connexions per no saturar la web
        time.sleep(sleeping_time)
    
    except Exception as e:
        
        # Si salta alguna excepció la mostrem per pantalla
        print(f"[ERROR] No es pot accedir a la URL: {i} / {total_urls} : {url}, excepció: {str(e)}")
        
# Mostrem per pantalla per quantes URLs de les totals s'ha pogut capturar tota la informació
print(f"\n URLs de pisos capturades: {total_urls} \n Nº de registres generats: {len(df_pisos)}")

print("Totes les URLS capturades be")

# Convertim la columna 'Any' en numèrica
df_pisos['Any'] = pd.to_numeric(df_pisos['Any'], errors='coerce')

# Creem una nova columna que calcula l'antiguitat del pis en anys
df_pisos['Temps'] = df_pisos['Any'].apply(lambda x: datetime.now().year - x if pd.notnull(x) else None)

# La insertem a continuació de la columna 'Any'
df_pisos.insert(12, 'Temps', df_pisos.pop('Temps'))

# Convertim les següents columnes en numèriques
df_pisos['Preu'] = df_pisos['Preu'].str.split(' ').str[0].str.replace('.', '', regex=False).astype(int)
df_pisos['Temps'] = pd.to_numeric(df_pisos['Temps'], errors='coerce')
df_pisos['m2_constr'] = df_pisos['m2_constr'].str.replace(' m2', '').astype(float)
df_pisos['m2_utils'] = df_pisos['m2_utils'].str.replace(' m2', '').astype(float)
df_pisos['Planta'] = pd.to_numeric(df_pisos['Planta'], errors='coerce')
df_pisos['Habitacions'] = pd.to_numeric(df_pisos['Habitacions'], errors='coerce')
df_pisos['Banys'] = pd.to_numeric(df_pisos['Banys'], errors='coerce')

# Afegim una nova columna de gap
df_pisos['Gap_2'] = ''

# Generem una nova columna que calcula el preu / m2
df_pisos['Preu/m2'] = df_pisos['Preu'] / df_pisos['m2_constr']

# Definim un coeficient de penalització (els valors poden anar entre 0.5% i 2%)
coef = 0.5 / 100

# Generem una nova columna que calcula el preu / m2 ponderat per l'any
df_pisos['Preu/m2/any'] = df_pisos['Preu/m2'] / (1 + (coef * df_pisos['Temps']))

# Finalment reemplaçem qualsevol valor NaN per guionets (-)
df_pisos.fillna('-', inplace=True)

print("A punt per cargar Excel")

# Definim una funció que calcula diferències de temps (en dies) entre dos dates
def days_between(d1, d2):
    d1 = datetime.strptime(d1, "%Y-%m-%d")
    d2 = datetime.strptime(d2, "%Y-%m-%d")
    return abs((d2 - d1).days)

# Definim la ruta on es troba l'arxiu Excel i el carreguem
ruta_excel = "New_home.xlsx"
book = load_workbook(ruta_excel)

print("Excel carregat")

# Definim la pestanya on bolcar la informació i la carreguem
hoja = 'SIP'
sheet = book[hoja]

# Definim la columna de l'Excel on es troben les referències dels pisos
columna_d = sheet['E']

# Fem un llistat de les referències associades a tots els pisos presents en l'Excel
referencias_excel = [cell.value for cell in columna_d if cell.value is not None]

# Fem un llistat de les referències associades a tots els pisos presents a la web de SIP
referencias_df = df_pisos['Referencia'].tolist()

# Comparem els dos llistats anteriors per detectar quins pisos ja han sigut venguts
referencias_vendidas = [ref for ref in referencias_excel[1:] if ref not in referencias_df]

# Fem un contador per anotar quants pisos nous s'han incorporat
total_nous = 0

# Fem un contador per anotar quants pisos existents simplement s'han actualitzat
total_actualitzats = 0

# Fem un contador per anotar quants pisos s'han vengut
total_venuts = 0

# Fem un contador per anotar quants pisos s'han reactivat
total_reactivats = 0

print("A punt de loopejar")

# Bloc de codi que actualitza l'estat de l'Excel dels pisos venuts
for fila in range(6, sheet.max_row):
    
    # Guardem la referència del pis associat a la fila que estem iterant
    referencia = sheet.cell(row=fila, column=5).value
    
    # Guardem l'estat del pis associat a la fila que estem iterant
    valor_columna_b = sheet.cell(row=fila, column=2).value

    if referencia in referencias_vendidas and valor_columna_b != 'Venut':

        # Si l'estat del pis és actiu però la referència apareix a la dels pisos venguts, el canviem a 'Venut'
        sheet.cell(row=fila, column=2).value = 'Venut'
        
        # Si es cumpleix la mateixa lògica anterior, també actualizem la data en que es marca com a 'Venut'
        sheet.cell(row=fila, column=6).value = datetime.today().strftime('%Y-%m-%d')
        
        # Finalment, calculem la diferència de temps entre que es va detectar el pis fins que s'ha venut
        valor_columna_c = sheet.cell(row=fila, column=3).value
        fecha_actual = datetime.today().strftime('%Y-%m-%d')
        actualizacion = days_between(str(fecha_actual), str(valor_columna_c))
        sheet.cell(row=fila, column=7).value = actualizacion
        
        # I actualitzem el contador de pisos venguts
        total_venuts += 1

print("Primer loop fet")

# Bloc de codi que actualitza la data dels pisos ja existents a l'Excel que segueixent apareixent a la web de SIP
for fila in range(6, sheet.max_row):
    
    # Guardem la referència del pis associat a la fila que estem iterant
    referencia = sheet.cell(row=fila, column=5).value
    
    # Guardem l'estat del pis associat a la fila que estem iterant    
    valor_columna_B = sheet.cell(row=fila, column=2).value

    if referencia in referencias_df and valor_columna_B != 'Venut':
        
        # Si l'estat del pis és actiu i la referència apareix al llistat de pisos obtinguts de la web de SIP, actualitzem la data
        sheet.cell(row=fila, column=6).value = datetime.today().strftime('%Y-%m-%d')

        # Finalment, calculem la diferència de temps entre que es va detectar el pis fins que s'ha actualitzat
        valor_columna_C = sheet.cell(row=fila, column=3).value
        fecha_actual = datetime.today().strftime('%Y-%m-%d')
        actualizacion = days_between(str(fecha_actual), str(valor_columna_C))
        sheet.cell(row=fila, column=7).value = actualizacion
        
        # I actualitzem el contador de pisos actualitzats
        total_actualitzats += 1

# Filtrem el dataframe de pisos capturats de la web de SIP per mantenir només aquells que han aparegut nous
df_nuevos_pisos = df_pisos[~df_pisos['Referencia'].isin(referencias_excel)]

print("Segon loop fet")

# Bloc de codi que apunta a l'Excel els pisos que han aparegut nous
if not df_nuevos_pisos.empty:
    
    # Definim la columna a partir de la qual es bolcarà la informació
    columna_b = sheet['B']
    
    # Busquem quina és la primera fila lliure de la columna definida anteriorment
    primera_fila_vacia = len([cell for cell in columna_b if cell.value is not None]) + 1

    # Afegim els pisos que han aparegut nous a l'Excel a partir de la primera fila lliure disponible
    with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='w', if_sheet_exists='replace') as writer:
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        df_nuevos_pisos.to_excel(writer, sheet_name=hoja, startrow=primera_fila_vacia, startcol=1, index=False, header=False)

    # Canviem les URLs afegides a l'Excel per hipervincles associats a la paraula 'Aqui' perquè sigui més visual
    for idx, url in enumerate(df_nuevos_pisos['URL'], start=primera_fila_vacia + 1):
        cell = sheet.cell(row=idx, column=29)
        cell.value = "Aqui"
        cell.hyperlink = url
        cell.style = "Hyperlink"

    # I actualitzem el contador de pisos actualitzats
    total_nous += len(df_nuevos_pisos)    

print("Tercer loop fet, previ a guardar excel")

# Guardem els canvis fets
book.save(ruta_excel)

print("Excel guardat")

# Afegim un últim bloc que comprova si un pis que estava marcat com a 'Venut' s'ha reactivat
if len(df_pisos) != total_nous + total_actualitzats + total_venuts:
    
    for fila in range(6, sheet.max_row):
    
        # Guardem la referència del pis associat a la fila que estem iterant
        referencia = sheet.cell(row=fila, column=5).value

        # Guardem l'estat del pis associat a la fila que estem iterant    
        valor_columna_B = sheet.cell(row=fila, column=2).value

        # Si el pis estava marcat en l'Excel com a 'Venut' però segueix apareixent a la web, vol dir que l'han reactivat
        if referencia in referencias_df and valor_columna_B == 'Venut':

            # Marquem els pisos que ja estaven marcats com a 'Venut' amb una nova referència '_1' per a diferenciar-los
            sheet.cell(row=fila, column=5).value = f"{referencia}_1"
            
            # Definim la columna a partir de la qual es bolcarà la nova informació
            columna_b = sheet['B']

            # Busquem quina és la primera fila lliure de la columna definida anteriorment
            primera_fila_vacia = len([cell for cell in columna_b if cell.value is not None]) + 1

            # Afegim els pisos reactivats amb les noves dades
            with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='w', if_sheet_exists='replace') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                df_pisos[df_pisos['Referencia'] == referencia].to_excel(writer, sheet_name=hoja, startrow=primera_fila_vacia, startcol=1, index=False, header=False)

            # Canviem les URLs afegides a l'Excel per hipervincles associats a la paraula 'Aqui' perquè sigui més visual
            for idx, url in enumerate(df_pisos[df_pisos['Referencia'] == referencia]['URL'], start=primera_fila_vacia + 1):
                cell = sheet.cell(row=idx, column=29)
                cell.value = "Aqui"
                cell.hyperlink = url
                cell.style = "Hyperlink"

            # I actualitzem el contador de pisos reactivats
            total_reactivats += 1

print("Ultim loop fet, previ a guardar de nou")

# Tornem a guardar els canvis fets
book.save(ruta_excel)

print("Ultima guardada normal")

# Calculem la data actual en format _YYYYMMDD
fecha_actual = datetime.now().strftime("_%Y%m%d")

# Fem una còpia de l'arxiu Excel afegint el sufix de la data
book.save(f"New_home{fecha_actual}.xlsx")
    
# Mostrem els resultats per pantalla
print(f" Dels {len(df_pisos)} pisos obtinguts de la web de SIP: \n\n  - {total_actualitzats} simplement s'han actualitzat a l'Excel. \n  - {total_nous} s'han afegit com a pisos nous a l'Excel. \n  - {total_reactivats} s'han reactivat a l'Excel. \n\n Per altra banda, en l'Excel s'han marcat {total_venuts} pisos ja existents com a 'Venuts'\n\n")
