#!/usr/bin/env python
# coding: utf-8

# Importem les llibreries necessàries
from openpyxl import load_workbook
from datetime import datetime
from bs4 import BeautifulSoup
import pandas as pd
import unicodedata
import requests
import time
import re

# Definim el nombre màxim de pàgines a consultar
max_pages = 2

# Definim el temps a esperar entre cada connexió a les diferents pàgines
sleeping_time = 15

# Definim una llista amb les URLs base que volem consultar
urls_base = ["https://www.serra.immo/buscador-de-inmuebles/page/{page}/?status%5B0%5D=comprar&type%5B0%5D=viviendas&location%5B0%5D=vilafranca-del-penedes&property_id&rooms&min-price&max-price",
             "https://www.serra.immo/buscador-de-inmuebles/page/{page}/?status%5B0%5D=comprar&type%5B0%5D=viviendas&location%5B0%5D=olerdola&property_id&rooms&min-price&max-price",
             "https://www.serra.immo/buscador-de-inmuebles/page/{page}/?status%5B0%5D=comprar&type%5B0%5D=viviendas&location%5B0%5D=vilobi-del-penedes&property_id&rooms&min-price&max-price",
             "https://www.serra.immo/buscador-de-inmuebles/page/{page}/?status%5B0%5D=comprar&type%5B0%5D=viviendas&location%5B0%5D=font-rubi&property_id&rooms&min-price&max-price",
             "https://www.serra.immo/buscador-de-inmuebles/page/{page}/?status%5B0%5D=comprar&type%5B0%5D=viviendas&location%5B0%5D=moja&property_id&rooms&min-price&max-price",
             "https://www.serra.immo/buscador-de-inmuebles/page/{page}/?status%5B0%5D=comprar&type%5B0%5D=viviendas&location%5B0%5D=les-cabanyes&property_id&rooms&min-price&max-price"]

# Llista per emmagatzemar les URLs dels pisos ofertats
urls_pisos = []

# Iterem sobre cada URL base
for base_url in urls_base:
    # Accedim a les N pàgines definides per cada URL base
    for page in range(1, max_pages + 1):
        # Construïm la URL completa
        url = base_url.format(page=page)
        
        try:
            # Fem la petició a la pàgina
            response = requests.get(url)
            
            # Verifiquem si la resposta ha estat exitosa
            if response.status_code == 200:
                print(f"[SUCCESS] Accedint correctament a {url} ...")
                
                # Capturem el contingut HTML amb BeautifulSoup
                soup = BeautifulSoup(response.content, 'html.parser')
                
                # Busquem totes les etiquetes 'a' de la classe que conté les URLs de cada pis
                links = soup.find_all('a', target='_self')
                
                # Obtenim les URLs de cada pis i les guardem a la variable 'urls_pisos'
                for link in links:
                    href = link.get('href')
                    if href:
                        urls_pisos.append(href)
                        
            else:
                print(f"[ERROR] No es pot accedir a {url}, estat: {response.status_code}")
                break
            
            # Esperem el temps d'espera entre connexions per no saturar la web
            time.sleep(sleeping_time)
        
        except Exception as e:
            print(f"[ERROR] No es pot accedir a {url}, excepció: {str(e)}")
            break

# Eliminem la meitat dels registres per evitar duplicats (si cal)
urls_pisos = urls_pisos[1::2]

# Guardem en una variable el total de URLs de pisos obtingudes
total_urls = len(urls_pisos)

# Mostrem per pantalla la quantitat total de URLs de pisos obtingudes
print(f"\nS'han pogut capturar {total_urls} URLs de pisos diferents")

# Generem un dataframe buit amb les columnes dels diferents atributs que volem capturar de cada URL
columnes = ['Estat', 'Data_captura', 'Plataforma', 'Referencia', 'Data_actualització', 'Diferencia_temps', 'Gap_1', 'Preu', 'Comunitat', 'Habitatge', 'Conservacio', 'Any', 'Ciutat', 'Barri', 'm2_constr', 'm2_utils', 'Planta', 'Ascensor', 'Habitacions', 'Banys', 'Orientacio', 'Jardi', 'Terrassa', 'Piscina', 'Garatge', 'Traster', 'URL']
df_pisos = pd.DataFrame(columns=columnes)

# Accedim a cada una de les URLs individuals de pisos capturades anteriorment
for i, url in enumerate(urls_pisos, start=1):
    
    try:
        # Fem la petició a la pàgina
        response = requests.get(url)
        
        # Verifiquem si ha sigut exitosa
        if response.status_code == 200:
            print(f"[SUCCESS] Accedint correctament a la URL {i} / {total_urls} : {url}")
            
            # Capturem el contignut HTML amb BeautifulSoup
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Extraiem el preu
            preu = soup.find('li', class_='item-price').get_text(strip=True)
            
            # Extraiem un primer llistat de característiques que no estan indexades com la resta
            altres = soup.find('ul', class_='list-3-cols list-unstyled')
            
            # Obtenim els valors del llistat anterior (en cas que existeixin)       
            try:
                altres_valors = [li.text.strip() for li in altres.find_all('li')]
            except Exception as e:
                altres_valors = []
            
            # Fem una llista dels atributs d'interès a buscar dins de la llista anterior
            ascensor = 'No'
            terrassa = 'No'
            
            # Accedim a la llista en busca dels atributs
            for texto in altres_valors:
                    
                if 'ascensor' in texto.lower(): 
                    ascensor = 'Si'
                    
                if 'terraza' in texto.lower(): 
                    terrassa = 'Si'
            
            # Capturem la descripció de l'habitatge
            descripcio = soup.find('div', class_='description-content').get_text(strip=True)
            
            # Fem una llista dels atributs d'interès a buscar dins de la descripció de l'habitatge
            jardi = 'No'
            piscina = 'No'
            traster = 'No'
            orientacio = '-'
            
            # Busquem si existeix la paraula 'jardin' a la descripció
            if any(word in unicodedata.normalize('NFD', descripcio).encode('ascii', 'ignore').decode('utf-8').lower() for word in ('jardin', 'jardi')):
                jardi = 'Si'            
            
            # Busquem si existeix la paraula 'piscina' a la descripció
            if 'piscina' in descripcio.lower(): 
                piscina = 'Si'

            # Busquem si existei la paraula 'trastero' a la descripció
            if any(word in descripcio.lower() for word in ('trastero', 'traster')):
                traster = 'Si'
            
            # Mirem si a la descripció es parla de l'orientació de l'habitatge
            match_orientacio = re.search(r'orientacion (\w+)|orientacio (\w+)', unicodedata.normalize('NFD', descripcio).encode('ascii', 'ignore').decode('utf-8'))

            # En cas afirmatiu, l'anotem
            if match_orientacio:
                orientacio = match_orientacio.group(1).title()
            
            # Extraiem un segon llistat de característiques que no estan indexades com la resta
            etiquetas = soup.find_all('li', class_='property-overview-item')
            
            # Fem una llista dels atributs d'interès a buscar dins de la llista anterior
            habitacions = '-'
            banys = '-'
            anys = '-'
            m2_constr = '-'
            garatge = 'No'

            # Accedim a la llista en busca dels atributs
            for indice, i in enumerate(etiquetas):
                if 'bed' in str(i):
                    habitacions = int(re.search(r'<strong>(.*?)</strong>', str(i)).group(1))

                if 'bathroom' in str(i):
                    banys = int(re.search(r'<strong>(.*?)</strong>', str(i)).group(1))

                if 'car' in str(i):
                    if int(re.search(r'<strong>(.*?)</strong>', str(i)).group(1)) > 0:
                        garatge = 'Si'

                if 'ruler' in str(i):
                    m2_constr = int(float(re.search(r'<strong>(.*?)</strong>', str(i)).group(1)))

                if 'calendar' in str(i):
                    anys = int(float(re.search(r'<strong>(.*?)</strong>', str(i)).group(1)))  
        
            # Definim un estat de conservació '-' en cas de que no es trobi
            conservacio = '-'

            # Busquem les etiquetes on surt l'estat de conservació
            label_conservacio = soup.find_all('div', class_='property-labels-wrap')

            # Ens assegurem que existeix un <div>
            if label_conservacio:

                # Agafem el primer element trobat
                label = label_conservacio[0]

                # Busquem tots els enllaços que conté
                links = label.find_all('a')

                # Ens quedem amb el que no es 'Comprar'
                found = False
                for link in links:
                    link_text = link.text.strip()

                    if link_text != 'Comprar':  
                        conservacio = link_text 
                        found = True
                        break 
            
            # Extraiem la resta de característiques que estan indexades correctament
            referencia = soup.find('strong', string='Referencia:').next_sibling.strip()
            comunitat = '-'
            habitatge = soup.find('li', class_='property-overview-item').get_text(strip=True)
            ciutat = soup.find('li', class_='detail-city').get_text(strip=True)[6:]
            barri = soup.find('li', class_='detail-area').get_text(strip=True)[6:]
            m2_utils = '-'
            planta = '-'
           
            # Afegim tota la informació capturada al dataframe creat anteriorment
            df_pisos = df_pisos.append({'Estat': 'Actiu',
                                        'Data_captura':datetime.today().strftime('%Y-%m-%d'),
                                        'Plataforma': 'Serra Grup Immobiliari',
                                        'Referencia': referencia,
                                        'Data_actualització':datetime.today().strftime('%Y-%m-%d'),
                                        'Diferencia_temps': 0,
                                        'Gap_1': '',
                                        'Preu': preu,
                                        'Comunitat': comunitat,
                                        'Habitatge': habitatge,
                                        'Conservacio': conservacio, 
                                        'Any': anys,
                                        'Ciutat': ciutat,
                                        'Barri': barri,
                                        'm2_constr': m2_constr,
                                        'm2_utils': m2_utils,
                                        'Planta': planta,
                                        'Ascensor': ascensor,
                                        'Habitacions': habitacions,
                                        'Banys': banys,
                                        'Orientacio': orientacio,
                                        'Jardi': jardi,
                                        'Terrassa': terrassa,
                                        'Piscina': piscina,
                                        'Garatge': garatge,
                                        'Traster': traster,
                                        'URL': url
                                       }, ignore_index=True)
        
        else:
            print(f"[ERROR] No es pot accedir a la URL {i} / {total_urls} : {url}, estat: {response.status_code}")

        # Esperem el temps d'espera entre connexions per no saturar la web
        time.sleep(sleeping_time)
    
    except Exception as e:
        
        # Si salta alguna excepció la mostrem per pantalla
        print(f"[ERROR] No es pot accedir a la URL: {i} / {total_urls} : {url}, excepció: {str(e)}")
        
# Mostrem per pantalla per quantes URLs de les totals s'ha pogut capturar tota la informació
print(f"\n URLs de pisos capturades: {total_urls} \n Nº de registres generats: {len(df_pisos)}")

# Convertim la columna 'Any' en numèrica
df_pisos['Any'] = pd.to_numeric(df_pisos['Any'], errors='coerce')

# Creem una nova columna que calcula l'antiguitat del pis en anys
df_pisos['Temps'] = df_pisos['Any'].apply(lambda x: datetime.now().year - x if pd.notnull(x) else None)

# La insertem a continuació de la columna 'Any'
df_pisos.insert(12, 'Temps', df_pisos.pop('Temps'))

# Convertim les següents columnes en numèriques
df_pisos['Preu'] = df_pisos['Preu'].str.split(' ').str[0].str.replace('.', '', regex=False).str[:-1].astype(int)
df_pisos['Temps'] = pd.to_numeric(df_pisos['Temps'], errors='coerce')
df_pisos['m2_constr'] = df_pisos['m2_constr'].astype(float)
df_pisos['Habitacions'] = pd.to_numeric(df_pisos['Habitacions'], errors='coerce')
df_pisos['Banys'] = pd.to_numeric(df_pisos['Banys'], errors='coerce')

# Afegim una nova columna de gap
df_pisos['Gap_2'] = ''

# Generem una nova columna que calcula el preu / m2
df_pisos['Preu/m2'] = df_pisos['Preu'] / df_pisos['m2_constr']

# Definim un coeficient de penalització (els valors poden anar entre 0.5% i 2%)
coef = 0.5 / 100

# Generem una nova columna que calcula el preu / m2 ponderat per l'any
df_pisos['Preu/m2/any'] = df_pisos['Preu/m2'] / (1 + (coef * df_pisos['Temps']))

# Finalment reemplaçem qualsevol valor NaN per guionets (-)
df_pisos.fillna('-', inplace=True)

# Definim una funció que calcula diferències de temps (en dies) entre dos dates
def days_between(d1, d2):
    d1 = datetime.strptime(d1, "%Y-%m-%d")
    d2 = datetime.strptime(d2, "%Y-%m-%d")
    return abs((d2 - d1).days)

# Definim la ruta on es troba l'arxiu Excel i el carreguem
ruta_excel = "New_home.xlsx"
book = load_workbook(ruta_excel)

# Definim la pestanya on bolcar la informació i la carreguem
hoja = 'SERRA'
sheet = book[hoja]

# Definim la columna de l'Excel on es troben les referències dels pisos
columna_d = sheet['E']

# Fem un llistat de les referències associades a tots els pisos presents en l'Excel
referencias_excel = [cell.value for cell in columna_d if cell.value is not None]

# Fem un llistat de les referències associades a tots els pisos presents a la web de SIP
referencias_df = df_pisos['Referencia'].tolist()

# Comparem els dos llistats anteriors per detectar quins pisos ja han sigut venguts
referencias_vendidas = [ref for ref in referencias_excel[1:] if ref not in referencias_df]

# Fem un contador per anotar quants pisos nous s'han incorporat
total_nous = 0

# Fem un contador per anotar quants pisos existents simplement s'han actualitzat
total_actualitzats = 0

# Fem un contador per anotar quants pisos s'han vengut
total_venuts = 0

# Fem un contador per anotar quants pisos s'han reactivat
total_reactivats = 0

# Bloc de codi que actualitza l'estat de l'Excel dels pisos venuts
for fila in range(6, sheet.max_row):
    
    # Guardem la referència del pis associat a la fila que estem iterant
    referencia = sheet.cell(row=fila, column=5).value
    
    # Guardem l'estat del pis associat a la fila que estem iterant
    valor_columna_b = sheet.cell(row=fila, column=2).value

    if referencia in referencias_vendidas and valor_columna_b != 'Venut':

        # Si l'estat del pis és actiu però la referència apareix a la dels pisos venguts, el canviem a 'Venut'
        sheet.cell(row=fila, column=2).value = 'Venut'
        
        # Si es cumpleix la mateixa lògica anterior, també actualizem la data en que es marca com a 'Venut'
        sheet.cell(row=fila, column=6).value = datetime.today().strftime('%Y-%m-%d')
        
        # Finalment, calculem la diferència de temps entre que es va detectar el pis fins que s'ha venut
        valor_columna_c = sheet.cell(row=fila, column=3).value
        fecha_actual = datetime.today().strftime('%Y-%m-%d')
        actualizacion = days_between(str(fecha_actual), str(valor_columna_c))
        sheet.cell(row=fila, column=7).value = actualizacion
        
        # I actualitzem el contador de pisos venguts
        total_venuts += 1

# Bloc de codi que actualitza la data dels pisos ja existents a l'Excel que segueixent apareixent a la web de SIP
for fila in range(6, sheet.max_row):
    
    # Guardem la referència del pis associat a la fila que estem iterant
    referencia = sheet.cell(row=fila, column=5).value
    
    # Guardem l'estat del pis associat a la fila que estem iterant    
    valor_columna_B = sheet.cell(row=fila, column=2).value

    if referencia in referencias_df and valor_columna_B != 'Venut':
        
        # Si l'estat del pis és actiu i la referència apareix al llistat de pisos obtinguts de la web de SIP, actualitzem la data
        sheet.cell(row=fila, column=6).value = datetime.today().strftime('%Y-%m-%d')

        # Finalment, calculem la diferència de temps entre que es va detectar el pis fins que s'ha actualitzat
        valor_columna_C = sheet.cell(row=fila, column=3).value
        fecha_actual = datetime.today().strftime('%Y-%m-%d')
        actualizacion = days_between(str(fecha_actual), str(valor_columna_C))
        sheet.cell(row=fila, column=7).value = actualizacion
        
        # I actualitzem el contador de pisos actualitzats
        total_actualitzats += 1

# Filtrem el dataframe de pisos capturats de la web de SIP per mantenir només aquells que han aparegut nous
df_nuevos_pisos = df_pisos[~df_pisos['Referencia'].isin(referencias_excel)]

# Bloc de codi que apunta a l'Excel els pisos que han aparegut nous
if not df_nuevos_pisos.empty:
    
    # Definim la columna a partir de la qual es bolcarà la informació
    columna_b = sheet['B']
    
    # Busquem quina és la primera fila lliure de la columna definida anteriorment
    primera_fila_vacia = len([cell for cell in columna_b if cell.value is not None]) + 1

    # Afegim els pisos que han aparegut nous a l'Excel a partir de la primera fila lliure disponible
    for row_index, row_data in enumerate(df_nuevos_pisos.values, start=primera_fila_vacia):
        for col_index, cell_value in enumerate(row_data, start=2): 
            sheet.cell(row=row_index, column=col_index, value=cell_value)

    # Canviem les URLs afegides a l'Excel per hipervincles associats a la paraula 'Aqui' perquè sigui més visual
    for idx, url in enumerate(df_nuevos_pisos['URL'], start=primera_fila_vacia):
        cell = sheet.cell(row=idx, column=29)
        cell.value = "Aqui"
        cell.hyperlink = url
        cell.style = "Hyperlink"

    # I actualitzem el contador de pisos actualitzats
    total_nous += len(df_nuevos_pisos)    

# Guardem els canvis fets
book.save(ruta_excel)

# Afegim un últim bloc que comprova si un pis que estava marcat com a 'Venut' s'ha reactivat
if len(df_pisos) != total_nous + total_actualitzats + total_venuts:
    
    for fila in range(6, sheet.max_row):
    
        # Guardem la referència del pis associat a la fila que estem iterant
        referencia = sheet.cell(row=fila, column=5).value

        # Guardem l'estat del pis associat a la fila que estem iterant    
        valor_columna_B = sheet.cell(row=fila, column=2).value

        # Si el pis estava marcat en l'Excel com a 'Venut' però segueix apareixent a la web, vol dir que l'han reactivat
        if referencia in referencias_df and valor_columna_B == 'Venut':

            # Marquem els pisos que ja estaven marcats com a 'Venut' amb una nova referència '_1' per a diferenciar-los
            sheet.cell(row=fila, column=5).value = f"{referencia}_1"
            
            # Definim la columna a partir de la qual es bolcarà la nova informació
            columna_b = sheet['B']

            # Busquem quina és la primera fila lliure de la columna definida anteriorment
            primera_fila_vacia = len([cell for cell in columna_b if cell.value is not None]) + 1

            # Afegim els pisos reactivats amb les noves dades
            for row_index, row_data in enumerate(df_pisos[df_pisos['Referencia'] == referencia].values, start=primera_fila_vacia):
                for col_index, cell_value in enumerate(row_data, start=2):  
                    sheet.cell(row=row_index, column=col_index, value=cell_value)

            # Canviem les URLs afegides a l'Excel per hipervincles associats a la paraula 'Aqui' perquè sigui més visual
            for idx, url in enumerate(df_pisos[df_pisos['Referencia'] == referencia]['URL'], start=primera_fila_vacia):
                cell = sheet.cell(row=idx, column=29)
                cell.value = "Aqui"
                cell.hyperlink = url
                cell.style = "Hyperlink"

            # I actualitzem el contador de pisos reactivats
            total_reactivats += 1

# Tornem a guardar els canvis fets
book.save(ruta_excel)

# Calculem la data actual en format _YYYYMMDD
fecha_actual = datetime.now().strftime("_%Y%m%d")

# Fem una còpia de l'arxiu Excel afegint el sufix de la data
book.save(f"New_home{fecha_actual}.xlsx")
    
# Mostrem els resultats per pantalla
print(f" Dels {len(df_pisos)} pisos obtinguts de la web de SIP: \n\n  - {total_actualitzats} simplement s'han actualitzat a l'Excel. \n  - {total_nous} s'han afegit com a pisos nous a l'Excel. \n  - {total_reactivats} s'han reactivat a l'Excel. \n\n Per altra banda, en l'Excel s'han marcat {total_venuts} pisos ja existents com a 'Venuts'\n\n")
